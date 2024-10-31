from typing import List
import pandas as pd
import os

from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def extract_specific_column(df: pd.DataFrame, column_name: str) -> List:
    # Remove leading/trailing whitespace
    df.columns = df.columns.str.strip()
    column_name = column_name.strip()

    if column_name in df.columns:
        return df[column_name].tolist()
    else:
        raise ValueError(f"Column '{column_name}' not found in the file.")

def path_modify(path: str, name: str, new: bool = False) -> str:
    path = os.path.normpath(path)
    filename = f"{name}_NEW.xlsx" if new else f"{name}.xlsx"
    return os.path.join(path, filename)

def extract_data(test_results_column: str,  file_name: str, path: str):
    file_path = path_modify(path, file_name)
    df = pd.read_excel(file_path)
    questions = extract_specific_column(df, "Questions")
    if not test_results_column:
        return df, questions, ""
    test_results = extract_specific_column(df, test_results_column)
    return df, questions, test_results


def create_new_file(df: pd.DataFrame, answers: list, file_name: str, path: str) -> None:
    df['New Answers'] = answers
    modified_path = path_modify(path, file_name, new=True)

    # Create a new Excel file with custom formatting
    from openpyxl import Workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Add the DataFrame data to the worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    # Set the width and background color of the 'New Answers' column
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    column_number_answers = len(df.columns)
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            worksheet.row_dimensions[row].height = 200
            worksheet.column_dimensions[get_column_letter(col)].width = 30

    workbook.save(modified_path)
    print(f"New file created: {modified_path}")




